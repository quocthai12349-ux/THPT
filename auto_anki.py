import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Alignment
import os
import re

# === CẤU HÌNH ===
folder_path = "."  # Thư mục chứa file PDF
excel_path = os.path.join(folder_path, "tracnghiem_tonghop.xlsx")

# === HÀM HỖ TRỢ ===
def clean_title(s, fallback):
    """Lấy tên bài rõ ràng, tránh Untitled."""
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    match = re.search(r"(Bài|Dạng|Bổ trợ)\s*\d*[:\-–]?\s*(.*)", s, re.IGNORECASE)
    if match:
        s = match.group(0)
    s = re.sub(r"(?i)\b(vật lý|physics|đề|trắc nghiệm|đúng sai)\b", "", s)
    s = s.strip(" -:").title()
    if not s or len(s) < 3:
        s = fallback.replace("(đề)", "").replace(".pdf", "").strip()
    return s


def format_question(text):
    """
    Giữ nguyên biểu thức, chỉ ngắt dòng đúng chỗ:
    - Xuống dòng trước A. B. C. D.
    - Xuống dòng trước a) b) c) d)
    Không ảnh hưởng đến công thức có 'a.' hoặc 'b.'.
    """
    # Nối từ bị ngắt dòng trong công thức
    text = text.replace("\n", " ")
    text = re.sub(r"(?<=\S)-\s+(?=\S)", "-", text)

    # Xuống dòng đúng chỗ: (A. B. C. D.) hoặc (a) b) c) d))
    text = re.sub(r"(?<![A-Za-z0-9])([A-D])\.", r"\n\1.", text)   # Trắc nghiệm A. B. C. D.
    text = re.sub(r"(?<![A-Za-z0-9])([a-d])\)", r"\n\1)", text)   # Đúng sai a) b) c) d)

    # Dọn khoảng trắng dư
    text = re.sub(r"[ \t]{2,}", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


# === XỬ LÝ PDF ===
def process_pdfs():
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
    if not pdf_files:
        print("⚠️ Không tìm thấy file PDF trong thư mục.")
        return []

    rows = []
    for pdf_name in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_name)
        doc = fitz.open(pdf_path)

        # Lấy tiêu đề bài từ phần đầu PDF
        first_page_text = doc.load_page(0).get_text("text")
        first_line = ""
        for ln in first_page_text.splitlines():
            if any(k in ln.lower() for k in ["bài", "dạng", "bổ trợ"]):
                first_line = ln
                break
        title = clean_title(first_line or first_page_text[:100], pdf_name)

        print(f"📘 Đang xử lý: {pdf_name}  →  {title}")

        for p in range(doc.page_count):
            page = doc.load_page(p)
            text = page.get_text("text")

            # Tách từng câu hỏi theo Câu 1, Câu 2...
            parts = re.split(r"(?i)(?=Câu\s*\d+[:.)])", text)
            for part in parts:
                part = part.strip()
                if len(part) < 10:
                    continue
                q_text = format_question(part)
                rows.append((q_text.strip(), title))
    return rows


# === XUẤT EXCEL ===
def export_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tracnghiem"
    ws["A1"] = "Câu hỏi"
    ws["C1"] = "Tên bài"

    for i, (q, t) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=3, value=t)
        ws.cell(row=i, column=1).alignment = Alignment(wrapText=True, vertical="top")
        ws.cell(row=i, column=3).alignment = Alignment(wrapText=True, vertical="top")

    wb.save(excel_path)
    print(f"\n✅ Đã tạo file Excel: {excel_path}")


# === MAIN ===
if __name__ == "__main__":
    rows = process_pdfs()
    if rows:
        export_excel(rows)
    else:
        print("Không có dữ liệu để ghi Excel.")
